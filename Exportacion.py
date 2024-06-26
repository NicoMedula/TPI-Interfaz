import tkinter as tk
from tkinter import *
from tkinter import ttk
from customtkinter import *
from PIL import Image, ImageTk
from tkinter import messagebox
import ast
from tkinter import ttk
from Pagina2 import *
import openpyxl as xl
import random

##############################################--------------------------------------

def Limpiar():
    nombreprod.delete(0,END)
    cantprod.delete(0,END)

##############################################--------------------------------------

def cargarPedido():

    precioFrutas={ "naranja":300, "pomelo":400, "limon":250, "mandarina":200 }
    precioAceites={"naranja":2000, "limon":2350, "mandarina":2100, "pomelo":2500}
    precioJugos={"naranja":1500, "limon":1800, "mandarina":1650, "pomelo":2100}

    archivo=xl.load_workbook("Data.xlsx")
    hoja1=archivo["Hoja1"]

    valorProd=nombreprod.get()
    valorCat=categorias.get()

    numerofila = hoja1.max_row + 1
    valorCantidad=int(cantprod.get())

    hoja1.cell(column=1,row=numerofila,value=valorProd)
    hoja1.cell(column=2,row=numerofila,value=valorCantidad)
    hoja1.cell(column=3,row=numerofila,value=valorCat)
    hoja1.cell(column=5,row=numerofila,value=numerofila-1)
    
    


    if valorCat == "Frutas" and valorProd.lower() in precioFrutas:
        precio_total = 0  

        precio_unitario = precioFrutas[valorProd.lower()]
        precio_total = valorCantidad * precio_unitario
        hoja1.cell(column=4, row=hoja1.max_row, value=precio_total)
   
    if valorCat == "Aceites" and valorProd.lower() in precioAceites:
        precio_total = 0  

        precio_unitario = precioAceites[valorProd.lower()]
        precio_total = valorCantidad * precio_unitario
        hoja1.cell(column=4, row=hoja1.max_row, value=precio_total)

    if valorCat == "Jugos" and valorProd.lower() in precioJugos:
        precio_total = 0  

        precio_unitario = precioJugos[valorProd.lower()]
        precio_total = valorCantidad * precio_unitario
        hoja1.cell(column=4, row=hoja1.max_row, value=precio_total)

    archivo.save("Data.xlsx")
    messagebox.showinfo(title="Cargado",message="Se cargo el pedido!")
    App7.destroy()
    

#######################################################################---------------------------------------

def mostrarPedidos():
    App = tk.Toplevel()
    App.geometry("925x500+300+200")
    App.configure(background="gray14")
    App.title("Mis Pedidos")

    Arbol = ttk.Treeview(App, columns=("Producto", "Cantidad", "Categoría","Precio","Nro pedido"), show='headings')
    Arbol.heading("Producto", text="Producto")
    Arbol.heading("Cantidad", text="Cantidad")
    Arbol.heading("Categoría", text="Categoría")
    Arbol.heading("Precio", text="Precio")
    Arbol.heading("Nro pedido", text="Nro de pedido")

    Arbol.pack(expand=True, fill='both')

    archivo = xl.load_workbook("Data.xlsx")
    hoja = archivo["Hoja1"]
    
    for row in hoja.iter_rows(min_row=2, values_only=True):  
        Arbol.insert("", tk.END, values=row)

##############################################--------------------------------------
def realizarpedido():
    global App7
    App7=tk.Toplevel()
    App7.geometry("400x400")
    App7.configure(background="gray14")
    App7.title("Exportaciones")

    img_path = "C:/Users/nicol/OneDrive/Escritorio/TPI Interfaz/Imagen/logo.png"
    img3 = Image.open(img_path)
    img3 = img3.resize((250, 250), Image.LANCZOS)
    photo_img3 = ImageTk.PhotoImage(img3)

    Labelimg3 = tk.Label(App7, image=photo_img3, border=0,bg="gray14")
    Labelimg3.image = photo_img3  
    Labelimg3.place(relx=0.2, rely=0.73, anchor="center")

    global nombreprod
    global categorias
    global cantprod

    categorias=CTkComboBox(App7,values=["Aceites","Frutas","Jugos"],fg_color="#FFA500",corner_radius=32,dropdown_fg_color="#FFA500")
    categorias.place(relx=0.6,rely=0.5, anchor="center")

    nombreprod=CTkEntry(master=App7, placeholder_text="Ingrese el producto", width=200,font=("bold",12)) 
    nombreprod.place(relx=0.6,rely=0.3, anchor="center")

    cantprod=CTkEntry(master=App7, placeholder_text="Ingrese cantidad", width=200,font=("bold",12)) 
    cantprod.place(relx=0.6,rely=0.4, anchor="center")

    Botonrealizar=(CTkButton(master=App7, text="Realizar pedido", corner_radius=32,fg_color="#FFA500",
                   hover_color="#FF4500",command=cargarPedido))
    Botonrealizar.place(relx=0.6,rely=0.65, anchor="center")

    BotonLimpiar=CTkButton(master=App7, text="Limpiar", corner_radius=32,fg_color="#FFA500",
                   hover_color="#FF4500",command=Limpiar)
    BotonLimpiar.place(relx=0.6,rely=0.75, anchor="center")

###############################################################------------------------------------------------------

def seguirmipedido():
    ventanaseguimiento = tk.Toplevel()
    ventanaseguimiento.geometry("925x500+300+200")
    ventanaseguimiento.configure(background="gray14")
    ventanaseguimiento.title("Seguir pedido")

    Arbol = ttk.Treeview(ventanaseguimiento, columns=("Producto", "Cantidad", "Estado", "Nro pedido"), show='headings')
    Arbol.heading("Producto", text="Producto")
    Arbol.heading("Cantidad", text="Cantidad")
    Arbol.heading("Estado", text="Estado")
    Arbol.heading("Nro pedido", text="Nro de pedido")
    Arbol.pack(expand=True, fill='both')

    archivo = xl.load_workbook("Data.xlsx")
    hoja1 = archivo["Hoja1"]

    for row in hoja1.iter_rows(min_row=2, values_only=True):
        producto, cantidad, categoria, _, numero_pedido = row 
        estado = random.choice(["En proceso", "En camino", "En espera"])  
        Arbol.insert("", tk.END, values=(producto, cantidad, estado, numero_pedido))

    archivo.save("Data.xlsx")


###############################################################------------------------------------------------------

def cuadroScroll():
    cuadro= CTkFrame(master=App4, fg_color="#8D6F3A", border_color="#FFCC70",border_width=2)
    cuadro.pack(expand=True)
    cuadro.place(rely=0.2,relx=0.5)
    
    CTkButton(master=cuadro, text="Mis pedidos",fg_color="#FFA500",hover_color="#FF4500",command=mostrarPedidos).pack(expand=True,padx=30,pady=20)               
    CTkButton(master=cuadro, text="Realizar pedido",fg_color="#FFA500",hover_color="#FF4500",command=realizarpedido).pack(expand=True,padx=30,pady=20)
    CTkButton(master=cuadro, text="Seguir mi pedido",fg_color="#FFA500",hover_color="#FF4500",command=seguirmipedido).pack(expand=True,padx=30,pady=20)
    
    headtexto=Label(App4,text="Secciones",fg="orange2",bg="gray14",font=("Microsoft Yahei UI Light",23,"bold"))
    headtexto.place(relx=0.55,rely=0.1)

def VentanaExport():
    global App4
    App4=tk.Toplevel()
    App4.geometry("925x500+300+200")
    App4.configure(background="gray14")
    App4.title("Exportaciones")
    cuadroScroll()


